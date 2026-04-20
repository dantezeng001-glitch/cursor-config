from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook


DEFAULT_CONFIG = {
    "source_patterns": ["*.xlsx", "*.xlsm", "**/*.xlsx", "**/*.xlsm"],
    "exclude_patterns": [
        "cursor_excel/**",
        ".cursor/**",
        ".vscode/**",
        "**/~$*.xlsx",
        "**/~$*.xlsm",
        "~$*.xlsx",
        "~$*.xlsm",
    ],
    "output_dir": "cursor_excel",
    "preview_rows": 20,
    "preview_columns": 12,
}

CONFIG_FILE_NAME = "excel_sync_config.json"


def load_config(project_root: Path) -> dict[str, Any]:
    config_path = project_root / CONFIG_FILE_NAME
    config = DEFAULT_CONFIG.copy()
    if config_path.exists():
        user_config = json.loads(config_path.read_text(encoding="utf-8"))
        config.update(user_config)
    return config


def is_excluded(relative_path: str, patterns: list[str]) -> bool:
    path = PurePosixPath(relative_path)
    return any(path.match(pattern) for pattern in patterns)


def discover_workbooks(project_root: Path, config: dict[str, Any]) -> list[Path]:
    candidates: set[Path] = set()
    for pattern in config["source_patterns"]:
        candidates.update(project_root.glob(pattern))

    workbooks: list[Path] = []
    for path in sorted(candidates):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        relative_path = path.relative_to(project_root).as_posix()
        if is_excluded(relative_path, config["exclude_patterns"]):
            continue
        if path.name.startswith("~$"):
            continue
        workbooks.append(path)
    return workbooks


def safe_name(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "untitled"


def workbook_id(project_root: Path, workbook_path: Path) -> str:
    relative_path = workbook_path.relative_to(project_root).as_posix()
    suffix = hashlib.md5(relative_path.encode("utf-8")).hexdigest()[:8]
    return f"{safe_name(workbook_path.stem)}_{suffix}"


def normalize_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def make_headers(first_row: list[Any], width: int) -> list[str]:
    used: dict[str, int] = {}
    headers: list[str] = []
    for index in range(width):
        raw = first_row[index] if index < len(first_row) else ""
        base = str(raw).strip() if raw not in ("", None) else f"column_{index + 1}"
        count = used.get(base, 0) + 1
        used[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def detect_header_row(rows: list[list[Any]], width: int) -> int:
    best_index = 0
    best_score = (-1, -1, -1.0)
    for index, row in enumerate(rows[:5]):
        padded = row + [""] * max(0, width - len(row))
        non_empty = [str(value).strip() for value in padded if str(value).strip()]
        if not non_empty:
            continue

        short_cells = sum(1 for value in non_empty if len(value) <= 40)
        unique_ratio = len(set(non_empty)) / len(non_empty)
        score = (len(non_empty), short_cells, unique_ratio)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def markdown_cell(value: Any) -> str:
    text = str(value) if value != "" else ""
    text = text.replace("\r\n", "<br>").replace("\n", "<br>")
    text = text.replace("|", "\\|")
    return text


def trim_row(row: list[Any]) -> list[Any]:
    items = list(row)
    while items and items[-1] == "":
        items.pop()
    return items


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        for row in rows:
            writer.writerow(row)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_markdown_preview(
    path: Path,
    workbook_name: str,
    workbook_relative_path: str,
    sheet_name: str,
    headers: list[str],
    leading_rows: list[list[Any]],
    data_rows: list[list[Any]],
    sheet_stats: dict[str, Any],
    preview_rows: int,
    preview_columns: int,
) -> None:
    lines = [
        f"# {workbook_name} / {sheet_name}",
        "",
        f"- Source workbook: `{workbook_relative_path}`",
        f"- Rows: `{sheet_stats['rows']}`",
        f"- Columns: `{sheet_stats['columns']}`",
        f"- Header row index: `{sheet_stats['header_row_index']}`",
        "",
        "## Columns",
        "",
        ", ".join(f"`{header}`" for header in headers) if headers else "No columns detected.",
        "",
        "## Preview",
        "",
    ]

    if leading_rows:
        lines.extend(
            [
                "## Notes Before Header",
                "",
                f"There are `{len(leading_rows)}` row(s) before the detected header. They are kept in the raw CSV and JSON output.",
                "",
            ]
        )

    if headers and data_rows:
        preview_headers = headers[:preview_columns]
        lines.append("| " + " | ".join(markdown_cell(item) for item in preview_headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(preview_headers)) + " |")
        for row in data_rows[:preview_rows]:
            clipped = row[:preview_columns]
            padded = clipped + [""] * max(0, len(preview_headers) - len(clipped))
            lines.append("| " + " | ".join(markdown_cell(item) for item in padded) + " |")
        if len(data_rows) > preview_rows:
            lines.extend(["", f"Only the first `{preview_rows}` rows are shown here. Use the sibling `.csv` or `.json` for the full sheet."])
    elif headers:
        lines.append("This sheet only contains a header row.")
    else:
        lines.append("This sheet is empty.")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def sync_workbook(
    project_root: Path,
    output_dir: Path,
    workbook_path: Path,
    config: dict[str, Any],
) -> dict[str, Any]:
    book_id = workbook_id(project_root, workbook_path)
    book_output_dir = output_dir / book_id
    if book_output_dir.exists():
        shutil.rmtree(book_output_dir)
    book_output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    workbook_relative_path = workbook_path.relative_to(project_root).as_posix()
    workbook_manifest: dict[str, Any] = {
        "id": book_id,
        "name": workbook_path.name,
        "source": workbook_relative_path,
        "updated_at": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(timespec="seconds"),
        "sheets": [],
    }

    for sheet in workbook.worksheets:
        rows: list[list[Any]] = []
        max_width = 0
        for row in sheet.iter_rows(values_only=True):
            normalized = trim_row([normalize_cell(value) for value in row])
            rows.append(normalized)
            max_width = max(max_width, len(normalized))

        sheet_file_stem = safe_name(sheet.title)
        csv_path = book_output_dir / f"{sheet_file_stem}.csv"
        json_path = book_output_dir / f"{sheet_file_stem}.json"
        md_path = book_output_dir / f"{sheet_file_stem}.md"

        write_csv(csv_path, rows)

        if rows and max_width:
            header_row_index = detect_header_row(rows, max_width)
            headers = make_headers(rows[header_row_index], max_width)
            leading_rows = rows[:header_row_index]
            data_rows = [row + [""] * (max_width - len(row)) for row in rows[header_row_index + 1 :]]
            records = [dict(zip(headers, row)) for row in data_rows]
        elif rows:
            headers = []
            header_row_index = 0
            leading_rows = []
            data_rows = []
            records = []
        else:
            headers = []
            header_row_index = 0
            leading_rows = []
            data_rows = []
            records = []

        sheet_payload = {
            "workbook": workbook_path.name,
            "sheet": sheet.title,
            "source": workbook_relative_path,
            "rows": len(rows),
            "columns": max_width,
            "header_row_index": header_row_index,
            "leading_rows": leading_rows,
            "headers": headers,
            "records": records,
        }
        write_json(json_path, sheet_payload)
        write_markdown_preview(
            md_path,
            workbook_path.name,
            workbook_relative_path,
            sheet.title,
            headers,
            leading_rows,
            data_rows,
            {"rows": len(rows), "columns": max_width, "header_row_index": header_row_index},
            config["preview_rows"],
            config["preview_columns"],
        )

        workbook_manifest["sheets"].append(
            {
                "name": sheet.title,
                "rows": len(rows),
                "columns": max_width,
                "header_row_index": header_row_index,
                "csv": csv_path.relative_to(project_root).as_posix(),
                "json": json_path.relative_to(project_root).as_posix(),
                "md": md_path.relative_to(project_root).as_posix(),
            }
        )

    write_json(book_output_dir / "manifest.json", workbook_manifest)
    workbook.close()
    return workbook_manifest


def cleanup_stale_outputs(output_dir: Path, active_ids: set[str]) -> None:
    if not output_dir.exists():
        return
    for child in output_dir.iterdir():
        if child.is_dir() and child.name not in active_ids:
            shutil.rmtree(child)


def build_index_markdown(index_path: Path, manifest: list[dict[str, Any]]) -> None:
    lines = [
        "# Excel Sync Index",
        "",
        "Read this file first when a request is about Excel, tables, campaign data, budgets, timelines, or tabular marketing content.",
        "",
        f"- Updated at: `{datetime.now().isoformat(timespec='seconds')}`",
        f"- Workbooks tracked: `{len(manifest)}`",
        "",
        "## Workbooks",
        "",
    ]

    if not manifest:
        lines.append("No Excel files are currently being tracked.")

    for workbook in manifest:
        lines.extend(
            [
                f"### {workbook['name']}",
                f"- Source: `{workbook['source']}`",
                f"- Updated at: `{workbook['updated_at']}`",
                f"- Workbook manifest: `cursor_excel/{workbook['id']}/manifest.json`",
            ]
        )
        for sheet in workbook["sheets"]:
            lines.append(
                f"- Sheet `{sheet['name']}` -> `{sheet['md']}`, `{sheet['csv']}`, `{sheet['json']}`"
            )
        lines.append("")

    index_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def run_sync(project_root: Path) -> list[dict[str, Any]]:
    config = load_config(project_root)
    output_dir = project_root / config["output_dir"]
    output_dir.mkdir(parents=True, exist_ok=True)

    workbooks = discover_workbooks(project_root, config)
    manifests = [sync_workbook(project_root, output_dir, workbook_path, config) for workbook_path in workbooks]
    cleanup_stale_outputs(output_dir, {item["id"] for item in manifests})

    index_json_path = output_dir / "index.json"
    index_md_path = output_dir / "index.md"
    write_json(index_json_path, manifests)
    build_index_markdown(index_md_path, manifests)
    return manifests


def _locate_md_cleanup_script() -> Path | None:
    """Best-effort lookup for the md-cleanup skill's cleanup script."""
    candidates: list[Path] = []
    cursor_home = os.environ.get("CURSOR_HOME")
    if cursor_home:
        candidates.append(Path(cursor_home) / "skills" / "md-cleanup" / "scripts" / "md_cleanup.py")
    candidates.append(Path.home() / ".cursor" / "skills" / "md-cleanup" / "scripts" / "md_cleanup.py")
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def run_md_cleanup(output_dir: Path) -> None:
    """Invoke the md-cleanup skill's script on every `.md` under output_dir.

    Best-effort: prints a diagnostic line if md-cleanup is missing, but never
    raises — the sync itself already succeeded.
    """
    script = _locate_md_cleanup_script()
    if script is None:
        print("  [md-cleanup] skill not found; skipping post-processing.")
        return

    md_files = sorted(str(p) for p in output_dir.rglob("*.md"))
    if not md_files:
        return

    try:
        subprocess.run(
            [sys.executable, str(script), *md_files, "--in-place"],
            check=False,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"  [md-cleanup] post-processing failed: {exc!r}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync Excel workbooks into Cursor-readable text files.")
    parser.add_argument(
        "--project-root",
        default=Path(__file__).resolve().parents[1],
        type=Path,
        help="Project root that contains the Excel files.",
    )
    parser.add_argument(
        "--no-cleanup",
        action="store_true",
        help="Skip the md-cleanup post-processing step.",
    )
    args = parser.parse_args()

    project_root = args.project_root.resolve()
    manifests = run_sync(project_root)
    print(f"Synced {len(manifests)} workbook(s) into cursor_excel.")

    if not args.no_cleanup:
        config = load_config(project_root)
        output_dir = project_root / config["output_dir"]
        run_md_cleanup(output_dir)


if __name__ == "__main__":
    main()
